import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


COLUNAS_CONTABIL = {
    "linha": "Cod.R",
    "cliente": "Nome da Conta",
    "valor": "S.Atual",
}

COLUNAS_CLIENTE = {
    "linha": None,
    "cliente": "nome",
    "valor": "valor",
}

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FONT_HEADER = Font(color="FFFFFF", bold=True)


def converter_valor(valor) -> float:
    if valor is None or pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return 0.0

    texto = texto.replace("R$", "").replace(" ", "")

    negativo = False
    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1]

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        valor_float = float(texto)
        if negativo:
            valor_float *= -1
        return valor_float
    except Exception:
        return 0.0


def localizar_coluna(df: pd.DataFrame, nome_esperado: str):
    if nome_esperado is None:
        return None

    if nome_esperado in df.columns:
        return nome_esperado

    for col in df.columns:
        if str(col).strip().lower() == str(nome_esperado).strip().lower():
            return col

    raise ValueError(f"Coluna '{nome_esperado}' não encontrada. Colunas disponíveis: {list(df.columns)}")


def ler_excel_seguro(caminho_arquivo: str) -> pd.DataFrame:
    ext = os.path.splitext(caminho_arquivo)[1].lower()

    try:
        if ext == ".xlsx":
            return pd.read_excel(caminho_arquivo, dtype=str, engine="openpyxl")
        return pd.read_excel(caminho_arquivo, dtype=str)
    except Exception as e:
        raise RuntimeError(f"Erro ao ler o arquivo '{caminho_arquivo}': {e}") from e


def formatar_aba(ws, moeda_cols=None, freeze="A2", auto_filter=True):
    if freeze:
        ws.freeze_panes = freeze

    if auto_filter and ws.max_row > 0 and ws.max_column > 0:
        ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            valor = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(valor))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    if moeda_cols:
        for col_name in moeda_cols:
            idx = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(1, c).value == col_name:
                    idx = c
                    break
            if idx:
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, idx).number_format = 'R$ #,##0.00'


def executar_reconciliacao(
    arquivo_contabil: str,
    arquivo_cliente: str,
    arquivo_saida: str,
    nome_empresa: str = "API"
):
    df_cont = ler_excel_seguro(arquivo_contabil)
    df_cli = ler_excel_seguro(arquivo_cliente)

    col_cont_cliente = localizar_coluna(df_cont, COLUNAS_CONTABIL["cliente"])
    col_cont_valor = localizar_coluna(df_cont, COLUNAS_CONTABIL["valor"])

    col_cli_cliente = localizar_coluna(df_cli, COLUNAS_CLIENTE["cliente"])
    col_cli_valor = localizar_coluna(df_cli, COLUNAS_CLIENTE["valor"])

    cont = df_cont[[col_cont_cliente, col_cont_valor]].copy()
    cont.columns = ["Cliente", "Saldo Contábil"]
    cont["Cliente"] = cont["Cliente"].fillna("").astype(str).str.strip()
    cont["Saldo Contábil"] = cont["Saldo Contábil"].apply(converter_valor)
    cont = cont.groupby("Cliente", as_index=False)["Saldo Contábil"].sum()

    cli = df_cli[[col_cli_cliente, col_cli_valor]].copy()
    cli.columns = ["Cliente Base", "Saldo Cliente"]
    cli["Cliente Base"] = cli["Cliente Base"].fillna("").astype(str).str.strip()
    cli["Saldo Cliente"] = cli["Saldo Cliente"].apply(converter_valor)
    cli = cli.groupby("Cliente Base", as_index=False)["Saldo Cliente"].sum()

    relatorio = cont.merge(
        cli,
        how="outer",
        left_on="Cliente",
        right_on="Cliente Base"
    )

    relatorio["Cliente"] = relatorio["Cliente"].fillna("")
    relatorio["Cliente Base"] = relatorio["Cliente Base"].fillna("")
    relatorio["Saldo Contábil"] = relatorio["Saldo Contábil"].fillna(0.0)
    relatorio["Saldo Cliente"] = relatorio["Saldo Cliente"].fillna(0.0)
    relatorio["Diferença"] = relatorio["Saldo Contábil"] - relatorio["Saldo Cliente"]

    resumo = pd.DataFrame({
        "Indicador": [
            "Empresa",
            "Data/Hora Processamento",
            "Total saldo contábil",
            "Total posição cliente",
            "Diferença conciliatória total",
        ],
        "Valor": [
            nome_empresa,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            relatorio["Saldo Contábil"].sum(),
            relatorio["Saldo Cliente"].sum(),
            relatorio["Diferença"].sum(),
        ]
    })

    pasta_saida = os.path.dirname(os.path.abspath(arquivo_saida))
    os.makedirs(pasta_saida, exist_ok=True)

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        relatorio.to_excel(writer, sheet_name="Relatorio", index=False)
        resumo.to_excel(writer, sheet_name="Resumo", index=False)

        ws_rel = writer.sheets["Relatorio"]
        formatar_aba(ws_rel, moeda_cols=["Saldo Contábil", "Saldo Cliente", "Diferença"])

        ws_res = writer.sheets["Resumo"]
        formatar_aba(ws_res)

    return arquivo_saida


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python reconciliador_api.py <arquivo_contabil> <arquivo_cliente> <arquivo_saida> [nome_empresa]")
        sys.exit(1)

    arquivo_contabil = sys.argv[1]
    arquivo_cliente = sys.argv[2]
    arquivo_saida = sys.argv[3]
    nome_empresa = sys.argv[4] if len(sys.argv) > 4 else "API"

    try:
        executar_reconciliacao(
            arquivo_contabil=arquivo_contabil,
            arquivo_cliente=arquivo_cliente,
            arquivo_saida=arquivo_saida,
            nome_empresa=nome_empresa
        )
        print(arquivo_saida)
        sys.exit(0)
    except Exception as e:
        print(f"ERRO: {e}")
        sys.exit(2)
import os
import sys
import json
import getpass
import socket
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


COLUNAS_CONTABIL = {
    "linha": "Cod.R",
    "fornecedor": "Nome da Conta",
    "valor": "S.Atual",
}

COLUNAS_FORNECEDOR = {
    "linha": None,
    "fornecedor": "nome",
    "valor": "valor",
}

TOP_DIVERGENCIAS = 100
MAX_ACESSOS_SALVAR = 300
NOME_ABA_PARETO = "Pareto"
NOME_ABA_ZERADOS = "Valores Zerados"
NOME_ABA_CORRECAO_CTB = "Possivel Correçao CTB"
ALTURA_LINHA_PADRAO = 20
ALTURA_CABECALHO = 40

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_CONCILIADO = PatternFill("solid", fgColor="E2F0D9")
FILL_DIVERGENCIA = PatternFill("solid", fgColor="FFF2CC")
FILL_SOMENTE = PatternFill("solid", fgColor="FCE4D6")
FILL_TOP = PatternFill("solid", fgColor="D9EAF7")
FILL_ZERADOS = PatternFill("solid", fgColor="EAF3FF")
FILL_CORRECAO = PatternFill("solid", fgColor="FDE9D9")
FONT_HEADER = Font(color="FFFFFF", bold=True)


def obter_area_trabalho() -> str:
    home = os.path.expanduser("~")
    for pasta in ("Desktop", "Área de Trabalho"):
        caminho = os.path.join(home, pasta)
        if os.path.exists(caminho):
            return caminho
    return home


def obter_pasta_logs() -> str:
    pasta = os.path.join(obter_area_trabalho(), "Logs_Conciliacao")
    os.makedirs(pasta, exist_ok=True)
    return pasta


def obter_caminho_log_acessos() -> str:
    return os.path.join(obter_pasta_logs(), "historico_acessos_conciliacao.json")


def normalizar_texto(texto) -> str:
    if texto is None or pd.isna(texto):
        return ""
    texto = str(texto).strip()
    if texto.lower() == "nan":
        return ""
    texto = texto.upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    texto = re.sub(r"[^A-Z0-9\s&/.\-]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def chave_reduzida(nome: str) -> str:
    nome = normalizar_texto(nome)
    stopwords = {
        "DA", "DE", "DI", "DO", "DU", "DAS", "DOS",
        "E", "LTDA", "ME", "EPP", "SA", "S", "A", "SOCIEDADE",
        "CLINICA", "ODONTOLOGICA", "ODONTOLOGICO",
    }
    partes = [p for p in nome.split() if p not in stopwords]
    return " ".join(partes).strip()


def converter_valor(valor) -> float:
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return 0.0

    texto = texto.replace("R$", "").replace(" ", "")
    negativo = False
    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1]

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        valor_float = float(texto)
        return -valor_float if negativo else valor_float
    except Exception:
        return 0.0


def juntar_linhas(series: pd.Series) -> str:
    linhas = []
    for item in series.dropna():
        texto = str(item).strip()
        if not texto or texto.lower() == "nan":
            continue
        try:
            linhas.append(str(int(float(item))))
        except Exception:
            linhas.append(texto)
    return ", ".join(dict.fromkeys(linhas))


def similaridade(a: str, b: str) -> float:
    return SequenceMatcher(None, chave_reduzida(a), chave_reduzida(b)).ratio()


def localizar_coluna(df: pd.DataFrame, nome_esperado: str):
    if nome_esperado is None:
        return None
    if nome_esperado in df.columns:
        return nome_esperado

    alvo = normalizar_texto(nome_esperado)
    for col in df.columns:
        if normalizar_texto(col) == alvo:
            return col

    raise ValueError(f"Coluna '{nome_esperado}' não encontrada. Colunas disponíveis: {list(df.columns)}")


def classificar_status(saldo_contabil: float, saldo_fornecedor: float, tolerancia: float) -> str:
    existe_contabil = abs(saldo_contabil) > tolerancia
    existe_fornecedor = abs(saldo_fornecedor) > tolerancia
    if existe_contabil and existe_fornecedor:
        if abs(saldo_contabil - saldo_fornecedor) <= tolerancia:
            return "Conciliado"
        return "Diferença de saldo"
    if existe_contabil and not existe_fornecedor:
        return "Somente contabilidade"
    if not existe_contabil and existe_fornecedor:
        return "Somente posição fornecedor"
    return "Sem movimento"


def ler_excel_seguro(caminho_arquivo: str) -> pd.DataFrame:
    ext = os.path.splitext(caminho_arquivo)[1].lower()
    try:
        if ext == ".xlsx":
            return pd.read_excel(caminho_arquivo, dtype=str, engine="openpyxl")
        return pd.read_excel(caminho_arquivo, dtype=str)
    except Exception as e:
        raise RuntimeError(f"Erro ao ler o arquivo '{caminho_arquivo}': {e}") from e


def preparar_base(caminho_arquivo: str, colunas_cfg: dict, tipo_base: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = ler_excel_seguro(caminho_arquivo)
    col_fornecedor = localizar_coluna(df, colunas_cfg["fornecedor"])
    col_valor = localizar_coluna(df, colunas_cfg["valor"])
    col_linha_cfg = colunas_cfg.get("linha")

    if col_linha_cfg:
      col_linha = localizar_coluna(df, colunas_cfg["linha"])
      base_original = df[[col_linha, col_fornecedor, col_valor]].copy()
      base_original.columns = ["linha_origem", "fornecedor_original", "valor"]
    else:
      base_original = df[[col_fornecedor, col_valor]].copy()
      base_original.columns = ["fornecedor_original", "valor"]
      base_original.insert(0, "linha_origem", "")

    base_original["fornecedor_original"] = base_original["fornecedor_original"].fillna("").astype(str).str.strip()
    base_original["linha_origem"] = base_original["linha_origem"].fillna("").astype(str).str.strip()
    base_original["fornecedor_normalizado"] = base_original["fornecedor_original"].apply(normalizar_texto)
    base_original["fornecedor_chave"] = base_original["fornecedor_original"].apply(chave_reduzida)
    base_original["valor"] = base_original["valor"].apply(converter_valor)
    base_original = base_original[(base_original["fornecedor_normalizado"] != "") | (base_original["valor"].abs() > 0)].copy()

    agrupado = (
        base_original.groupby("fornecedor_chave", dropna=False)
        .agg(
            Fornecedor=("fornecedor_normalizado", lambda x: x.dropna().iloc[0] if not x.dropna().empty else ""),
            Fornecedor_Original=("fornecedor_original", lambda x: x.dropna().iloc[0] if not x.dropna().empty else ""),
            Fornecedor_Normalizado=("fornecedor_normalizado", lambda x: x.dropna().iloc[0] if not x.dropna().empty else ""),
            Fornecedor_Chave=("fornecedor_chave", lambda x: x.dropna().iloc[0] if not x.dropna().empty else ""),
            Saldo=("valor", "sum"),
            Linhas=("linha_origem", juntar_linhas),
            Qtde_Linhas=("linha_origem", lambda x: x.astype(str).str.strip().replace("", pd.NA).dropna().shape[0]),
        )
        .reset_index(drop=True)
    )
    agrupado["Tipo Base"] = tipo_base
    return base_original, agrupado


def montar_registro_relatorio(row_cont=None, row_forn=None, similaridade_fuzzy=None):
    saldo_cont = float(row_cont["Saldo Contábil"]) if row_cont is not None else 0.0
    saldo_forn = float(row_forn["Saldo Fornecedor"]) if row_forn is not None else 0.0

    fornecedor_exibicao = ""
    if row_cont is not None and str(row_cont.get("Fornecedor", "")).strip():
        fornecedor_exibicao = row_cont["Fornecedor"]
    elif row_forn is not None and str(row_forn.get("Fornecedor Base", "")).strip():
        fornecedor_exibicao = row_forn["Fornecedor Base"]

    return {
        "Fornecedor": fornecedor_exibicao,
        "Código Contabilidade": row_cont["Código Contabilidade"] if row_cont is not None else "",
        "Saldo Contábil": saldo_cont,
        "Fornecedor Base": row_forn["Fornecedor Base"] if row_forn is not None else "",
        "Saldo Fornecedor": saldo_forn,
        "Diferença": saldo_cont - saldo_forn,
        "Similaridade Fuzzy": similaridade_fuzzy if similaridade_fuzzy is not None else "",
    }


def conciliar_bases_com_fuzzy(contabil: pd.DataFrame, fornecedor: pd.DataFrame, usar_fuzzy: bool, limiar_fuzzy: float) -> pd.DataFrame:
    contabil = contabil.copy().reset_index(drop=True)
    fornecedor = fornecedor.copy().reset_index(drop=True)
    contabil["__id_cont"] = contabil.index
    fornecedor["__id_forn"] = fornecedor.index

    registros = []
    ids_cont_usados = set()
    ids_forn_usados = set()

    exato_chave = contabil.merge(fornecedor, how="inner", on="Fornecedor_Chave", suffixes=("_cont", "_forn"))
    for _, row in exato_chave.iterrows():
        if row["__id_cont"] in ids_cont_usados or row["__id_forn"] in ids_forn_usados:
            continue
        row_cont = {"Fornecedor": row["Fornecedor_cont"], "Código Contabilidade": row["Código Contabilidade"], "Saldo Contábil": row["Saldo Contábil"]}
        row_forn = {"Fornecedor Base": row["Fornecedor Base"], "Saldo Fornecedor": row["Saldo Fornecedor"]}
        registros.append(montar_registro_relatorio(row_cont, row_forn, "1.0000"))
        ids_cont_usados.add(row["__id_cont"])
        ids_forn_usados.add(row["__id_forn"])

    cont_rest = contabil[~contabil["__id_cont"].isin(ids_cont_usados)].copy()
    forn_rest = fornecedor[~fornecedor["__id_forn"].isin(ids_forn_usados)].copy()

    if not cont_rest.empty and not forn_rest.empty:
        exato_normalizado = cont_rest.merge(forn_rest, how="inner", left_on="Fornecedor_Normalizado", right_on="Fornecedor_Normalizado", suffixes=("_cont", "_forn"))
        for _, row in exato_normalizado.iterrows():
            if row["__id_cont"] in ids_cont_usados or row["__id_forn"] in ids_forn_usados:
                continue
            row_cont = {"Fornecedor": row["Fornecedor_cont"], "Código Contabilidade": row["Código Contabilidade"], "Saldo Contábil": row["Saldo Contábil"]}
            row_forn = {"Fornecedor Base": row["Fornecedor Base"], "Saldo Fornecedor": row["Saldo Fornecedor"]}
            registros.append(montar_registro_relatorio(row_cont, row_forn, "1.0000"))
            ids_cont_usados.add(row["__id_cont"])
            ids_forn_usados.add(row["__id_forn"])

    cont_rest = contabil[~contabil["__id_cont"].isin(ids_cont_usados)].copy()
    forn_rest = fornecedor[~fornecedor["__id_forn"].isin(ids_forn_usados)].copy()

    if usar_fuzzy and not cont_rest.empty and not forn_rest.empty:
        pares_fuzzy = []
        for _, rc in cont_rest.iterrows():
            for _, rf in forn_rest.iterrows():
                score = similaridade(rc["Fornecedor"], rf["Fornecedor Base"])
                if score >= limiar_fuzzy:
                    pares_fuzzy.append({"__id_cont": rc["__id_cont"], "__id_forn": rf["__id_forn"], "score": score})

        pares_fuzzy = sorted(pares_fuzzy, key=lambda x: x["score"], reverse=True)
        cont_pareados = set()
        forn_pareados = set()

        for par in pares_fuzzy:
            if par["__id_cont"] in cont_pareados or par["__id_forn"] in forn_pareados:
                continue
            rc = cont_rest.loc[cont_rest["__id_cont"] == par["__id_cont"]].iloc[0]
            rf = forn_rest.loc[forn_rest["__id_forn"] == par["__id_forn"]].iloc[0]
            row_cont = {"Fornecedor": rc["Fornecedor"], "Código Contabilidade": rc["Código Contabilidade"], "Saldo Contábil": rc["Saldo Contábil"]}
            row_forn = {"Fornecedor Base": rf["Fornecedor Base"], "Saldo Fornecedor": rf["Saldo Fornecedor"]}
            registros.append(montar_registro_relatorio(row_cont, row_forn, f"{par['score']:.4f}"))
            cont_pareados.add(par["__id_cont"])
            forn_pareados.add(par["__id_forn"])

        cont_rest = cont_rest[~cont_rest["__id_cont"].isin(cont_pareados)].copy()
        forn_rest = forn_rest[~forn_rest["__id_forn"].isin(forn_pareados)].copy()

    for _, rc in cont_rest.iterrows():
        registros.append(montar_registro_relatorio(row_cont={"Fornecedor": rc["Fornecedor"], "Código Contabilidade": rc["Código Contabilidade"], "Saldo Contábil": rc["Saldo Contábil"]}))
    for _, rf in forn_rest.iterrows():
        registros.append(montar_registro_relatorio(row_forn={"Fornecedor Base": rf["Fornecedor Base"], "Saldo Fornecedor": rf["Saldo Fornecedor"]}))

    return pd.DataFrame(registros)


def gerar_fuzzy_matches(relatorio_base: pd.DataFrame) -> pd.DataFrame:
    if relatorio_base.empty:
        return pd.DataFrame(columns=["Fornecedor Contabilidade", "Codigo Contabilidade", "Saldo Contábil", "Fornecedor Base", "Saldo Fornecedor", "Diferença", "Similaridade"])

    fuzzy = relatorio_base[(relatorio_base["Similaridade Fuzzy"] != "") & (relatorio_base["Similaridade Fuzzy"] != "1.0000")].copy()
    if fuzzy.empty:
        return pd.DataFrame(columns=["Fornecedor Contabilidade", "Codigo Contabilidade", "Saldo Contábil", "Fornecedor Base", "Saldo Fornecedor", "Diferença", "Similaridade"])

    fuzzy["Similaridade"] = pd.to_numeric(fuzzy["Similaridade Fuzzy"], errors="coerce")
    fuzzy = fuzzy.rename(columns={"Fornecedor": "Fornecedor Contabilidade", "Código Contabilidade": "Codigo Contabilidade"})
    fuzzy = fuzzy[["Fornecedor Contabilidade", "Codigo Contabilidade", "Saldo Contábil", "Fornecedor Base", "Saldo Fornecedor", "Diferença", "Similaridade"]].copy()
    return fuzzy.sort_values(["Similaridade", "Fornecedor Contabilidade"], ascending=[False, True])


def adicionar_linha_subtotal(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame([{"Fornecedor": "SUBTOTAL", "Código Contabilidade": "", "Saldo Contábil": 0.0, "Fornecedor Base": "", "Saldo Fornecedor": 0.0, "Diferença": 0.0, "Status": "", "Similaridade Fuzzy": ""}])

    subtotal_linha = pd.DataFrame([{
        "Fornecedor": "SUBTOTAL",
        "Código Contabilidade": "",
        "Saldo Contábil": df["Saldo Contábil"].sum() if "Saldo Contábil" in df.columns else 0.0,
        "Fornecedor Base": "",
        "Saldo Fornecedor": df["Saldo Fornecedor"].sum() if "Saldo Fornecedor" in df.columns else 0.0,
        "Diferença": df["Diferença"].sum() if "Diferença" in df.columns else 0.0,
        "Status": "",
        "Similaridade Fuzzy": "",
    }])
    return pd.concat([df, subtotal_linha], ignore_index=True)


def montar_sheet_possivel_correcao_ctb(df_somente_contabilidade: pd.DataFrame) -> pd.DataFrame:
    if df_somente_contabilidade is None or df_somente_contabilidade.empty:
        return pd.DataFrame(columns=["Fornecedor", "Código Contabilidade", "Data", "Debito", "Credito", "Historico", "Saldo"])

    base = df_somente_contabilidade.copy()
    base = base[base["Fornecedor"] != "SUBTOTAL"].copy()
    if base.empty:
        return pd.DataFrame(columns=["Fornecedor", "Código Contabilidade", "Data", "Debito", "Credito", "Historico", "Saldo"])

    base["Data"] = ""
    base["Saldo"] = base["Diferença"]
    base["Debito"] = ""
    base["Credito"] = ""
    base["Historico"] = base["Fornecedor"].fillna("").astype(str).str.strip().apply(lambda nome: f"Baixa de saldos conforme alinhamento e comprovação da Administração - {nome}".strip())
    base.loc[base["Saldo"] > 0, "Credito"] = base.loc[base["Saldo"] > 0, "Código Contabilidade"]
    base.loc[base["Saldo"] < 0, "Debito"] = base.loc[base["Saldo"] < 0, "Código Contabilidade"]
    return base[["Fornecedor", "Código Contabilidade", "Data", "Debito", "Credito", "Historico", "Saldo"]].copy()


def formatar_aba(ws, moeda_cols=None, freeze="A2", auto_filter=True):
    if freeze:
        ws.freeze_panes = freeze
    if auto_filter and ws.max_row > 0 and ws.max_column > 0:
        ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 1:
        ws.row_dimensions[1].height = ALTURA_CABECALHO
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = ALTURA_LINHA_PADRAO

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                valor = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(valor))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 80)

    if moeda_cols:
        for col_name in moeda_cols:
            idx = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(1, c).value == col_name:
                    idx = c
                    break
            if idx:
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, idx).number_format = "#,##0.00"


def aplicar_cores_status(ws, nome_col_status="Status"):
    idx_status = None
    idx_fornecedor = None
    for c in range(1, ws.max_column + 1):
        valor_header = ws.cell(1, c).value
        if valor_header == nome_col_status:
            idx_status = c
        if valor_header == "Fornecedor":
            idx_fornecedor = c

    for r in range(2, ws.max_row + 1):
        fornecedor_valor = ws.cell(r, idx_fornecedor).value if idx_fornecedor else None
        status = ws.cell(r, idx_status).value if idx_status else None

        if fornecedor_valor == "SUBTOTAL":
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = FILL_HEADER
                ws.cell(r, c).font = FONT_HEADER
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            continue

        if status == "Conciliado":
            fill = FILL_CONCILIADO
        elif status == "Diferença de saldo":
            fill = FILL_DIVERGENCIA
        elif status in ("Somente contabilidade", "Somente posição fornecedor"):
            fill = FILL_SOMENTE
        else:
            fill = None

        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill


def aplicar_preenchimento(ws, fill, ignorar_subtotal=True):
    for r in range(2, ws.max_row + 1):
        if ignorar_subtotal and ws.cell(r, 1).value == "SUBTOTAL":
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = FILL_HEADER
                ws.cell(r, c).font = FONT_HEADER
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill


def ocultar_colunas(ws, colunas):
    for col in colunas:
        ws.column_dimensions[col].hidden = True


def ocultar_sheet(ws):
    ws.sheet_state = "hidden"


def obter_usuario_sistema() -> str:
    try:
        usuario = getpass.getuser()
        if usuario:
            return usuario
    except Exception:
        pass
    try:
        usuario = os.environ.get("USERNAME") or os.environ.get("USER")
        if usuario:
            return usuario
    except Exception:
        pass
    return "DESCONHECIDO"


def obter_nome_maquina() -> str:
    try:
        return socket.gethostname()
    except Exception:
        return "DESCONHECIDA"


def carregar_historico_acessos() -> list:
    caminho = obter_caminho_log_acessos()
    if not os.path.exists(caminho):
        return []
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            dados = json.load(f)
            return dados if isinstance(dados, list) else []
    except Exception:
        return []


def salvar_historico_acessos(registros: list):
    caminho = obter_caminho_log_acessos()
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(registros[-MAX_ACESSOS_SALVAR:], f, ensure_ascii=False, indent=2)


def registrar_acesso(tipo_acesso: str, nome_informado: str = "", empresa: str = "", arquivo_base: str = ""):
    registros = carregar_historico_acessos()
    registros.append({
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "tipo_acesso": str(tipo_acesso or "").strip(),
        "usuario_windows": obter_usuario_sistema(),
        "maquina": obter_nome_maquina(),
        "nome_informado": str(nome_informado or "").strip(),
        "empresa": str(empresa or "").strip(),
        "arquivo_base": os.path.basename(arquivo_base) if arquivo_base else "",
    })
    salvar_historico_acessos(registros)


def executar_conciliacao(arquivo_contabil: str, arquivo_fornecedor: str, arquivo_saida: str, nome_empresa: str, usar_fuzzy: bool, limiar_fuzzy: float, tolerancia_valor: float, progresso_callback=None, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    def progresso(valor):
        if progresso_callback:
            progresso_callback(valor)

    inicio = datetime.now()
    progresso(5)
    log("Iniciando processamento...")
    os.makedirs(os.path.dirname(os.path.abspath(arquivo_saida)), exist_ok=True)

    log("Lendo base contábil...")
    _, contabil = preparar_base(arquivo_contabil, COLUNAS_CONTABIL, "Contabilidade")
    contabil = contabil.rename(columns={"Saldo": "Saldo Contábil", "Linhas": "Código Contabilidade", "Qtde_Linhas": "Qtde Códigos Contabilidade", "Fornecedor_Original": "Nome Contábil Original"})
    progresso(20)

    log("Lendo base do fornecedor...")
    _, fornecedor = preparar_base(arquivo_fornecedor, COLUNAS_FORNECEDOR, "Posição Fornecedor")
    fornecedor = fornecedor.rename(columns={"Saldo": "Saldo Fornecedor", "Fornecedor_Original": "Fornecedor Base"})
    progresso(35)

    log("Executando cruzamento das bases...")
    relatorio = conciliar_bases_com_fuzzy(contabil=contabil, fornecedor=fornecedor, usar_fuzzy=usar_fuzzy, limiar_fuzzy=limiar_fuzzy)
    relatorio["Status"] = relatorio.apply(lambda x: classificar_status(x["Saldo Contábil"], x["Saldo Fornecedor"], tolerancia_valor), axis=1)
    relatorio = relatorio[["Fornecedor", "Código Contabilidade", "Saldo Contábil", "Fornecedor Base", "Saldo Fornecedor", "Diferença", "Status", "Similaridade Fuzzy"]].copy()
    relatorio = relatorio.sort_values(["Status", "Fornecedor", "Fornecedor Base"], ascending=[True, True, True]).reset_index(drop=True)

    total_contabil = float(relatorio["Saldo Contábil"].sum())
    total_fornecedor = float(relatorio["Saldo Fornecedor"].sum())
    diferenca_total = float(relatorio["Diferença"].sum())
    qtd_conciliado = int((relatorio["Status"] == "Conciliado").sum())
    qtd_divergencias = int((relatorio["Status"] != "Conciliado").sum())
    progresso(55)

    resumo = pd.DataFrame({
        "Indicador": ["Empresa", "Data/Hora Processamento", "Total saldo contábil", "Total posição fornecedor", "Diferença conciliatória total", "Quantidade conciliados", "Quantidade divergências", "Tolerância aplicada", "Fuzzy match ativado", "Limiar fuzzy"],
        "Valor": [nome_empresa or "", datetime.now().strftime("%d/%m/%Y %H:%M:%S"), total_contabil, total_fornecedor, diferenca_total, qtd_conciliado, qtd_divergencias, tolerancia_valor, "Sim" if usar_fuzzy else "Não", limiar_fuzzy if usar_fuzzy else ""],
    })

    resumo_status = relatorio["Status"].value_counts().rename_axis("Status").reset_index(name="Quantidade")
    top_div = relatorio[relatorio["Status"] == "Diferença de saldo"].copy()
    if not top_div.empty:
        top_div["Diferença Absoluta"] = top_div["Diferença"].abs()
        top_div = top_div.sort_values("Diferença Absoluta", ascending=False).head(TOP_DIVERGENCIAS).drop(columns=["Diferença Absoluta"])

    somente_cont = relatorio[relatorio["Status"] == "Somente contabilidade"].copy()
    somente_forn = relatorio[relatorio["Status"] == "Somente posição fornecedor"].copy()
    valores_zerados = relatorio[relatorio["Diferença"].abs() <= tolerancia_valor].copy()
    possivel_correcao_ctb = montar_sheet_possivel_correcao_ctb(somente_cont)
    fuzzy_df = gerar_fuzzy_matches(relatorio)

    relatorio = adicionar_linha_subtotal(relatorio)
    somente_cont = adicionar_linha_subtotal(somente_cont)
    somente_forn = adicionar_linha_subtotal(somente_forn)
    top_div = adicionar_linha_subtotal(top_div)
    valores_zerados = adicionar_linha_subtotal(valores_zerados)
    progresso(70)

    log("Gerando arquivo Excel...")
    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        relatorio.to_excel(writer, sheet_name="Relatorio", index=False)
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        resumo_status.to_excel(writer, sheet_name="Resumo Status", index=False)
        somente_cont.to_excel(writer, sheet_name="Somente Contabilidade", index=False)
        somente_forn.to_excel(writer, sheet_name="Somente Posicao Fornecedor", index=False)
        top_div.to_excel(writer, sheet_name=NOME_ABA_PARETO, index=False)
        valores_zerados.to_excel(writer, sheet_name=NOME_ABA_ZERADOS, index=False)
        possivel_correcao_ctb.to_excel(writer, sheet_name=NOME_ABA_CORRECAO_CTB, index=False)
        if usar_fuzzy:
            fuzzy_df.to_excel(writer, sheet_name="Possiveis Matches", index=False)

        ws_rel = writer.sheets["Relatorio"]
        formatar_aba(ws_rel, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
        aplicar_cores_status(ws_rel)
        ws_res = writer.sheets["Resumo"]
        formatar_aba(ws_res, moeda_cols=["Valor"])
        ws_res_status = writer.sheets["Resumo Status"]
        formatar_aba(ws_res_status, moeda_cols=["Quantidade"])
        ws_sc = writer.sheets["Somente Contabilidade"]
        formatar_aba(ws_sc, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
        aplicar_cores_status(ws_sc)
        ws_sf = writer.sheets["Somente Posicao Fornecedor"]
        formatar_aba(ws_sf, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
        aplicar_cores_status(ws_sf)
        ws_top = writer.sheets[NOME_ABA_PARETO]
        formatar_aba(ws_top, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
        aplicar_preenchimento(ws_top, FILL_TOP)
        ws_zer = writer.sheets[NOME_ABA_ZERADOS]
        formatar_aba(ws_zer, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
        aplicar_preenchimento(ws_zer, FILL_ZERADOS)
        ws_corr = writer.sheets[NOME_ABA_CORRECAO_CTB]
        formatar_aba(ws_corr, moeda_cols=["Saldo"])
        aplicar_preenchimento(ws_corr, FILL_CORRECAO, ignorar_subtotal=False)
        ocultar_colunas(ws_corr, ["A", "B"])

        if usar_fuzzy and "Possiveis Matches" in writer.sheets:
            ws_fuzzy = writer.sheets["Possiveis Matches"]
            formatar_aba(ws_fuzzy, moeda_cols=["Saldo Contábil", "Saldo Fornecedor", "Diferença"])
            for c in range(1, ws_fuzzy.max_column + 1):
                if ws_fuzzy.cell(1, c).value == "Similaridade":
                    for r in range(2, ws_fuzzy.max_row + 1):
                        ws_fuzzy.cell(r, c).number_format = "0.00%"

        ocultar_sheet(writer.sheets["Resumo Status"])
        ocultar_sheet(writer.sheets["Somente Contabilidade"])
        ocultar_sheet(writer.sheets["Somente Posicao Fornecedor"])
        if usar_fuzzy and "Possiveis Matches" in writer.sheets:
            ocultar_sheet(writer.sheets["Possiveis Matches"])

    progresso(100)
    tempo = datetime.now() - inicio
    log("")
    log("=" * 70)
    log("CONCILIAÇÃO CONCLUÍDA COM SUCESSO")
    log(f"Empresa: {nome_empresa or 'Não informada'}")
    log(f"Arquivo gerado: {arquivo_saida}")
    log(f"Total saldo contábil: {total_contabil:,.2f}")
    log(f"Total posição fornecedor: {total_fornecedor:,.2f}")
    log(f"Diferença conciliatória total: {diferenca_total:,.2f}")
    log(f"Quantidade conciliados: {qtd_conciliado}")
    log(f"Quantidade divergências: {qtd_divergencias}")
    log(f"Quantidade valores zerados: {len(valores_zerados) - 1 if not valores_zerados.empty else 0}")
    log(f"Quantidade possível correção CTB: {len(possivel_correcao_ctb)}")
    if usar_fuzzy:
        log(f"Possíveis matches fuzzy: {len(fuzzy_df)}")
    log(f"Log de acessos: {obter_caminho_log_acessos()}")
    log(f"Tempo total: {tempo}")
    log("=" * 70)


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python reconciliador_fornecedores_api.py <arquivo_contabil> <arquivo_fornecedor> <arquivo_saida> [nome_empresa] [usar_fuzzy] [limiar_fuzzy] [tolerancia]")
        sys.exit(1)

    arquivo_contabil = sys.argv[1]
    arquivo_fornecedor = sys.argv[2]
    arquivo_saida = sys.argv[3]
    nome_empresa = sys.argv[4] if len(sys.argv) > 4 else "API"
    usar_fuzzy = str(sys.argv[5]).strip().lower() != "false" if len(sys.argv) > 5 else True
    limiar_fuzzy = float(sys.argv[6]) if len(sys.argv) > 6 else 0.88
    tolerancia = float(sys.argv[7]) if len(sys.argv) > 7 else 0.01

    try:
        registrar_acesso("EXECUCAO_API", "API", nome_empresa, arquivo_contabil)
        executar_conciliacao(
            arquivo_contabil=arquivo_contabil,
            arquivo_fornecedor=arquivo_fornecedor,
            arquivo_saida=arquivo_saida,
            nome_empresa=nome_empresa,
            usar_fuzzy=usar_fuzzy,
            limiar_fuzzy=limiar_fuzzy,
            tolerancia_valor=tolerancia,
            progresso_callback=None,
            log_callback=print,
        )
        print(arquivo_saida)
        sys.exit(0)
    except Exception as e:
        print(f"ERRO: {e}")
        sys.exit(2)
